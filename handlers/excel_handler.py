from handlers.file_handler import FileHandler

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_CO_SAN = True
except ImportError:
    OPENPYXL_CO_SAN = False


class ExcelHandler(FileHandler):
    MAU_HEADER = "4472C4"
    MAU_HANG_CHAN = "DCE6F1"

    def __init__(self, duong_dan: str):
        if not duong_dan.endswith(".xlsx"):
            duong_dan += ".xlsx"
        super().__init__(duong_dan)
        if not OPENPYXL_CO_SAN:
            raise ImportError("Cần cài openpyxl: pip install openpyxl")

    def doc(self) -> list[dict]:
        if not self.ton_tai():
            return []
        try:
            wb = openpyxl.load_workbook(self.duong_dan)
            ws = wb.active
            du_lieu = []
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                ban_ghi = {}
                for header, gia_tri in zip(headers, row):
                    ban_ghi[header] = gia_tri
                du_lieu.append(ban_ghi)
            return du_lieu
        except Exception as e:
            raise IOError(f"Không thể đọc file Excel: {self.duong_dan}\n{e}")

    def ghi(self, du_lieu: list[dict]) -> None:
        if not du_lieu:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = list(du_lieu[0].keys())

            # --- Ghi và định dạng header ---
            header_fill = PatternFill("solid", fgColor=self.MAU_HEADER)
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for col, ten_cot in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=ten_cot)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            chan_fill = PatternFill("solid", fgColor=self.MAU_HANG_CHAN)
            for row_idx, ban_ghi in enumerate(du_lieu, start=2):
                fill = chan_fill if row_idx % 2 == 0 else PatternFill()
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=ban_ghi.get(header))
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left")

            for col in ws.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(self.duong_dan)
        except Exception as e:
            raise IOError(f"Không thể ghi file Excel: {self.duong_dan}\n{e}")